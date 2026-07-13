"""
Levisol Monthly Planning Engine.
Reads 'Levisol Planning Tool - INPUTS.xlsx', validates it, recomputes inventory norms,
derives the January net requirement, solves the production & distribution MILP, and writes
'Levisol Plan - OUTPUT.xlsx'. Designed to NEVER crash on bad input: every failure path ends
with a readable Run Log sheet and a console message telling the planner what to fix.

Usage: py planning_engine.py [inputs.xlsx] [output.xlsx]
"""
import os
import sys
import time
import traceback
import pandas as pd
import numpy as np
from scipy.stats import norm as normal_dist
import pulp
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

TOOL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUTS = sys.argv[1] if len(sys.argv) > 1 else os.path.join(TOOL_DIR, "Levisol Planning Tool - INPUTS.xlsx")
OUTPUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(TOOL_DIR, "Levisol Plan - OUTPUT.xlsx")

MONTHS = [m + " (kL)" for m in ["Jul-25", "Aug-25", "Sep-25", "Oct-25", "Nov-25", "Dec-25"]]
CAP_LINES = ["<=1.5 LT", "3-5 LT", "7-20 LT", "50 LT", "180-210 LT"]
CAP_COLS = {ln: f"Capacity {ln} (kL)" for ln in CAP_LINES}
HUBS = ["MHW", "MHE"]

run_log = []          # (level, message) - level in {"INFO","WARN","ERROR"}
def log(level, msg):
    run_log.append((level, msg))
    print(f"  [{level}] {msg}")

# ------------------------------------------------------------------ output helpers
def style_sheet(ws, df):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for idx, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(55, len(str(col)) + 4))

def write_output(sheets):
    """sheets: list of (name, DataFrame). Always includes Run Log. Handles the
    'output file is open in Excel' case by falling back to a timestamped name."""
    log_df = pd.DataFrame(run_log, columns=["Level", "Message"])
    all_sheets = [("Run Log", log_df)] + sheets
    target = OUTPUT
    try:
        writer = pd.ExcelWriter(target, engine="openpyxl")
    except PermissionError:
        target = OUTPUT.replace(".xlsx", time.strftime(" %Y-%m-%d %H%M%S") + ".xlsx")
        print(f"\n  NOTE: '{os.path.basename(OUTPUT)}' is open in Excel - writing to "
              f"'{os.path.basename(target)}' instead.")
        writer = pd.ExcelWriter(target, engine="openpyxl")
    with writer:
        for name, df in all_sheets:
            df.to_excel(writer, sheet_name=name[:31], index=False)
            style_sheet(writer.sheets[name[:31]], df)
    print(f"\n  Output written to: {target}")
    return target

def fail(msg):
    log("ERROR", msg)
    log("ERROR", "Run stopped. Fix the input above, save & close the INPUTS workbook, and run again.")
    write_output([])
    print("\n  RUN FAILED - see messages above and the Run Log sheet in the output file.")
    sys.exit(1)

# ------------------------------------------------------------------ 1. LOAD + VALIDATE
print("\nSTEP 1/5  Reading and validating inputs...")
if not os.path.exists(INPUTS):
    print(f"  ERROR: cannot find '{INPUTS}'. It must sit in the Planning Tool folder.")
    sys.exit(1)

REQUIRED = {
    "Settings": ["Setting", "Value"],
    "Jan Demand Forecast": ["Product Name", "CFA", "Jan-26 Forecast (kL)"],
    "Plants & Capacity": ["Plant Code", "Production Cost (Rs per kL)"] + list(CAP_COLS.values()),
    "Transport Plant-Hub": ["From Plant Code", "To MHW (Rs per kL)", "To MHE (Rs per kL)"],
    "Transport Hub-CFA": ["CFA", "From MHW (Rs per kL)", "From MHE (Rs per kL)"],
    "SKU Master": ["Product Name", "Capacity Line", "Penalty Cost (Rs per kL)", "Contractual (Yes/No)"],
    "Service Levels (Tiers)": ["Tier", "Volume Share", "Target Fill Rate"],
    "Sales History": ["Product Name", "CFA"] + MONTHS,
    "Lead Times": ["Product Name", "CFA", "Source Hub", "Plant to Hub LT (days)", "Hub to CFA LT (days)",
                    "Production LT (days)", "Production Variability (days)", "Transit Variability (days)"],
    "Opening Inventory CFA": ["Product Name", "CFA", "Opening Inventory (kL)"],
    "Opening Inventory Hub": ["Product Name", "Hub", "Opening Inventory (kL)"],
    "Hub SS Override": ["Product Name", "Hub (MHW/MHE)", "Override Safety Stock (kL)"],
}
try:
    xls = pd.ExcelFile(INPUTS)
except PermissionError:
    print("  ERROR: the INPUTS workbook is open in Excel. Close it and run again.")
    sys.exit(1)

data = {}
for sheet, cols in REQUIRED.items():
    if sheet not in xls.sheet_names:
        fail(f"Sheet '{sheet}' is missing from the INPUTS workbook.")
    df = pd.read_excel(xls, sheet_name=sheet)
    missing = [c for c in cols if c not in df.columns]
    if missing:
        fail(f"Sheet '{sheet}' is missing column(s): {missing}. Restore the header row exactly.")
    data[sheet] = df

settings_raw = dict(zip(data["Settings"]["Setting"].astype(str).str.strip(), data["Settings"]["Value"]))
def get_setting(name, cast, default, lo=None, hi=None):
    try:
        v = cast(settings_raw.get(name, default))
    except (ValueError, TypeError):
        fail(f"Settings: '{name}' must be a number - found '{settings_raw.get(name)}'.")
    if lo is not None and (v < lo or (hi is not None and v > hi)):
        fail(f"Settings: '{name}' = {v} is outside the sensible range [{lo}, {hi}].")
    return v

TIME_LIMIT = int(get_setting("Solver time limit (seconds)", float, 150, 10, 3600))
CONTR_MULT = get_setting("Contractual penalty multiplier", float, 5, 1, 100)
HUB_SHORT_FACTOR = get_setting("Hub shortfall penalty factor", float, 0.5, 0, 10)
BATCH = get_setting("Batch size (kL)", float, 25, 1, 1000)
HUB_SL = get_setting("Hub service level", float, 0.98, 0.5, 0.999)
WDAYS = get_setting("Working days per month", float, 30, 1, 31)
REQ_BASIS = str(settings_raw.get("Requirement basis (ROP or SS)", "ROP")).strip().upper()
if REQ_BASIS not in ("ROP", "SS"):
    fail(f"Settings: 'Requirement basis (ROP or SS)' must be ROP or SS - found '{REQ_BASIS}'.")
FORCE_CONTR = str(settings_raw.get("Force contractual supply (Yes/No)", "No")).strip().lower() == "yes"

plants_df = data["Plants & Capacity"].dropna(subset=["Plant Code"])
PLANTS = plants_df["Plant Code"].astype(str).str.strip().tolist()
if len(PLANTS) == 0:
    fail("Plants & Capacity: no plants found.")
for col in list(CAP_COLS.values()) + ["Production Cost (Rs per kL)"]:
    bad = plants_df[pd.to_numeric(plants_df[col], errors="coerce").isna() | (pd.to_numeric(plants_df[col], errors="coerce") < 0)]
    if len(bad):
        fail(f"Plants & Capacity: column '{col}' has a negative or non-numeric value for plant(s) {bad['Plant Code'].tolist()}.")

sku_df = data["SKU Master"].dropna(subset=["Product Name"])
SKUS = sku_df["Product Name"].astype(str).str.strip().tolist()
bad_line = sku_df[~sku_df["Capacity Line"].isin(CAP_LINES)]
if len(bad_line):
    fail(f"SKU Master: 'Capacity Line' must be one of {CAP_LINES}. Bad rows: {bad_line['Product Name'].tolist()[:5]}")
if (pd.to_numeric(sku_df["Penalty Cost (Rs per kL)"], errors="coerce").fillna(-1) < 0).any():
    fail("SKU Master: 'Penalty Cost (Rs per kL)' contains negative or non-numeric values.")

tp_hc = data["Transport Hub-CFA"].dropna(subset=["CFA"])
CFAS = tp_hc["CFA"].astype(str).str.strip().tolist()
for col in ["From MHW (Rs per kL)", "From MHE (Rs per kL)"]:
    if (pd.to_numeric(tp_hc[col], errors="coerce").fillna(-1) < 0).any():
        fail(f"Transport Hub-CFA: '{col}' contains negative or non-numeric values.")
tp_ph = data["Transport Plant-Hub"].dropna(subset=["From Plant Code"])
for col in ["To MHW (Rs per kL)", "To MHE (Rs per kL)"]:
    if (pd.to_numeric(tp_ph[col], errors="coerce").fillna(-1) < 0).any():
        fail(f"Transport Plant-Hub: '{col}' contains negative or non-numeric values.")
missing_plants = set(PLANTS) - set(tp_ph["From Plant Code"].astype(str).str.strip())
if missing_plants:
    fail(f"Transport Plant-Hub: no freight rates for plant(s) {sorted(missing_plants)}.")

demand = data["Jan Demand Forecast"].dropna(subset=["Product Name", "CFA"]).copy()
demand["Product Name"] = demand["Product Name"].astype(str).str.strip()
demand["CFA"] = demand["CFA"].astype(str).str.strip()
demand["Jan-26 Forecast (kL)"] = pd.to_numeric(demand["Jan-26 Forecast (kL)"], errors="coerce")
if demand["Jan-26 Forecast (kL)"].isna().any():
    bad = demand[demand["Jan-26 Forecast (kL)"].isna()]
    fail(f"Jan Demand Forecast: non-numeric forecast for e.g. {bad[['Product Name','CFA']].values[:3].tolist()}.")
neg = demand[demand["Jan-26 Forecast (kL)"] < 0]
if len(neg):
    fail(f"Jan Demand Forecast: negative forecast for {neg[['Product Name','CFA']].values[:3].tolist()}. Use 0 for no demand.")
unknown_sku = set(demand["Product Name"]) - set(SKUS)
if unknown_sku:
    fail(f"Jan Demand Forecast: product(s) not in SKU Master: {sorted(unknown_sku)[:5]}. Add them to SKU Master or fix the spelling.")
unknown_cfa = set(demand["CFA"]) - set(CFAS)
if unknown_cfa:
    fail(f"Jan Demand Forecast: CFA(s) with no freight rates in Transport Hub-CFA: {sorted(unknown_cfa)[:5]}.")
dupes = demand.duplicated(subset=["Product Name", "CFA"]).sum()
if dupes:
    fail(f"Jan Demand Forecast: {dupes} duplicated (Product, CFA) rows - each combination may appear only once.")

zero_lanes = int((demand["Jan-26 Forecast (kL)"] == 0).sum())
if zero_lanes:
    log("INFO", f"{zero_lanes} demand lanes have zero forecast - handled (no production planned for them).")
log("INFO", f"Inputs valid: {len(SKUS)} SKUs, {len(PLANTS)} plants, {len(CFAS)} CFAs, {len(demand)} demand lanes.")
log("INFO", f"Settings: time limit {TIME_LIMIT}s, contractual x{CONTR_MULT:g}, hub shortfall factor {HUB_SHORT_FACTOR:g}, "
             f"basis {REQ_BASIS}, force contractual: {'Yes' if FORCE_CONTR else 'No'}.")

# ------------------------------------------------------------------ 2. NORMS
print("STEP 2/5  Computing inventory norms...")
hist = data["Sales History"].dropna(subset=["Product Name", "CFA"]).copy()
for c in MONTHS:
    hist[c] = pd.to_numeric(hist[c], errors="coerce").fillna(0.0)
lt = data["Lead Times"].dropna(subset=["Product Name", "CFA"]).copy()
bad_hub = lt[~lt["Source Hub"].isin(HUBS)]
if len(bad_hub):
    fail(f"Lead Times: 'Source Hub' must be MHW or MHE. Bad rows: {bad_hub[['Product Name','CFA']].values[:3].tolist()}")

# tiers from volume shares (editable)
tiers_df = data["Service Levels (Tiers)"].dropna(subset=["Tier"])
tiers_df = tiers_df.sort_values("Volume Share", ascending=False)
cum_threshold, thresholds = 0.0, []
for _, r in tiers_df.iterrows():
    cum_threshold += float(r["Volume Share"])
    thresholds.append((cum_threshold, str(r["Tier"]).strip(), float(r["Target Fill Rate"])))
vol = hist.groupby("Product Name")[MONTHS].sum().sum(axis=1).sort_values(ascending=False)
cum = vol.cumsum() / max(vol.sum(), 1e-9)
def to_tier(cp):
    for thr, tier_name, _ in thresholds:
        if cp <= thr + 1e-9:
            return tier_name
    return thresholds[-1][1]
sku_tier = cum.apply(to_tier).to_dict()
tier_fill = {t: fr for _, t, fr in thresholds}
tier_z = {t: float(normal_dist.ppf(fr)) for t, fr in tier_fill.items()}

norms = hist.merge(lt, on=["Product Name", "CFA"], how="inner")
n_no_lt = len(hist) - len(norms)
if n_no_lt:
    log("WARN", f"{n_no_lt} sales-history lanes have no Lead Times row - norms not computed for them.")
norms["mu_d"] = norms[MONTHS].mean(axis=1) / WDAYS
norms["sd_d"] = norms[MONTHS].std(axis=1, ddof=1) / np.sqrt(WDAYS)
norms["LT"] = norms["Production LT (days)"] + norms["Plant to Hub LT (days)"] + norms["Hub to CFA LT (days)"]
norms["sLT"] = np.sqrt(norms["Production Variability (days)"]**2 + norms["Transit Variability (days)"]**2)
norms["Tier"] = norms["Product Name"].map(sku_tier).fillna(thresholds[-1][1])
norms["Z"] = norms["Tier"].map(tier_z)
norms["Safety Stock (kL)"] = norms["Z"] * np.sqrt(norms["LT"] * norms["sd_d"]**2 + norms["mu_d"]**2 * norms["sLT"]**2)
norms["Reorder Point (kL)"] = norms["mu_d"] * norms["LT"] + norms["Safety Stock (kL)"]
norms["Days of Cover"] = np.where(norms["mu_d"] > 0, norms["Reorder Point (kL)"] / norms["mu_d"], 0.0)

# hub norms (flat service level) with demand-weighted lead-time collapse + overrides
HUB_Z = float(normal_dist.ppf(HUB_SL))
leg_tot = norms["Plant to Hub LT (days)"] + norms["Hub to CFA LT (days)"]
norms["ph_var"] = norms["Transit Variability (days)"] * np.where(leg_tot > 0, norms["Plant to Hub LT (days)"] / leg_tot, 0.5)
hub_rows = []
for (sku, hub), grp in norms.groupby(["Product Name", "Source Hub"]):
    w = grp["mu_d"]
    wa = (lambda v: float(np.average(v, weights=w)) if w.sum() > 0 else float(v.mean()))
    mu = float(w.sum())
    sd = float(np.sqrt((grp["sd_d"]**2).sum()))
    lt_avg = wa(grp["Production LT (days)"]) + float(grp["Plant to Hub LT (days)"].mean())
    lt_sd = float(np.sqrt(wa(grp["Production Variability (days)"])**2 + wa(grp["ph_var"])**2))
    ss = HUB_Z * np.sqrt(lt_avg * sd**2 + mu**2 * lt_sd**2)
    hub_rows.append({"Product Name": sku, "Hub": hub, "Hub Avg Daily Demand (kL)": mu,
                     "Hub Safety Stock (kL)": ss,
                     "Hub Reorder Point (kL)": mu * lt_avg + ss,
                     "Hub Days of Cover": (mu * lt_avg + ss) / mu if mu > 0 else 0.0})
hub_norms = pd.DataFrame(hub_rows)

ovr = data["Hub SS Override"].dropna(subset=["Product Name"])
n_ovr = 0
for _, r in ovr.iterrows():
    m = (hub_norms["Product Name"] == str(r["Product Name"]).strip()) & (hub_norms["Hub"] == str(r["Hub (MHW/MHE)"]).strip())
    v = pd.to_numeric(r["Override Safety Stock (kL)"], errors="coerce")
    if pd.isna(v) or v < 0:
        log("WARN", f"Hub SS Override: ignored invalid override for {r['Product Name']} ({r['Hub (MHW/MHE)']}).")
        continue
    if m.any():
        hub_norms.loc[m, "Hub Safety Stock (kL)"] = float(v)
        n_ovr += 1
    else:
        hub_norms = pd.concat([hub_norms, pd.DataFrame([{
            "Product Name": str(r["Product Name"]).strip(), "Hub": str(r["Hub (MHW/MHE)"]).strip(),
            "Hub Avg Daily Demand (kL)": 0.0, "Hub Safety Stock (kL)": float(v),
            "Hub Reorder Point (kL)": float(v), "Hub Days of Cover": 0.0}])], ignore_index=True)
        n_ovr += 1
if n_ovr:
    log("INFO", f"Applied {n_ovr} hub safety-stock override(s).")
log("INFO", f"Norms computed: {len(norms)} SKU-CFA rows, {len(hub_norms)} SKU-hub rows.")

# ------------------------------------------------------------------ 3. NET REQUIREMENT
print("STEP 3/5  Deriving January net requirement...")
opc = data["Opening Inventory CFA"].dropna(subset=["Product Name", "CFA"]).copy()
opc["Opening Inventory (kL)"] = pd.to_numeric(opc["Opening Inventory (kL)"], errors="coerce").fillna(0.0)
oph = data["Opening Inventory Hub"].dropna(subset=["Product Name", "Hub"]).copy()
oph["Opening Inventory (kL)"] = pd.to_numeric(oph["Opening Inventory (kL)"], errors="coerce").fillna(0.0)

req = demand.merge(norms[["Product Name", "CFA", "Tier", "Safety Stock (kL)", "Reorder Point (kL)", "Source Hub"]],
                   on=["Product Name", "CFA"], how="left")
n_new = int(req["Reorder Point (kL)"].isna().sum())
if n_new:
    log("WARN", f"{n_new} demand lanes have no sales history/lead time (new lanes?) - norms treated as 0, "
                 f"natural hub defaulted to MHW for them.")
req["Safety Stock (kL)"] = req["Safety Stock (kL)"].fillna(0.0)
req["Reorder Point (kL)"] = req["Reorder Point (kL)"].fillna(0.0)
req["Source Hub"] = req["Source Hub"].fillna("MHW")
req["Tier"] = req["Tier"].fillna(thresholds[-1][1])
req = req.merge(opc, on=["Product Name", "CFA"], how="left")
req["Opening Inventory (kL)"] = req["Opening Inventory (kL)"].fillna(0.0)
buffer_col = "Reorder Point (kL)" if REQ_BASIS == "ROP" else "Safety Stock (kL)"
req["Net Requirement (kL)"] = (req["Jan-26 Forecast (kL)"] + req[buffer_col] - req["Opening Inventory (kL)"]).clip(lower=0)
log("INFO", f"Total net requirement ({REQ_BASIS} basis): {req['Net Requirement (kL)'].sum():,.1f} kL "
             f"across {int((req['Net Requirement (kL)'] > 0).sum())} lanes.")

# ------------------------------------------------------------------ 4. OPTIMISE
print(f"STEP 4/5  Optimising (up to {TIME_LIMIT}s)...")
prod_cost = dict(zip(plants_df["Plant Code"].astype(str).str.strip(), plants_df["Production Cost (Rs per kL)"]))
cap = {(p, ln): float(plants_df.loc[plants_df["Plant Code"].astype(str).str.strip() == p, CAP_COLS[ln]].values[0])
       for p in PLANTS for ln in CAP_LINES}
ph_rate = {}
for _, r in tp_ph.iterrows():
    p = str(r["From Plant Code"]).strip()
    ph_rate[(p, "MHW")] = float(r["To MHW (Rs per kL)"])
    ph_rate[(p, "MHE")] = float(r["To MHE (Rs per kL)"])
hc_rate = {}
for _, r in tp_hc.iterrows():
    cfa = str(r["CFA"]).strip()
    hc_rate[("MHW", cfa)] = float(r["From MHW (Rs per kL)"])
    hc_rate[("MHE", cfa)] = float(r["From MHE (Rs per kL)"])
sku_line = dict(zip(sku_df["Product Name"].astype(str).str.strip(), sku_df["Capacity Line"]))
sku_pen = dict(zip(sku_df["Product Name"].astype(str).str.strip(), pd.to_numeric(sku_df["Penalty Cost (Rs per kL)"])))
sku_contr = {str(r["Product Name"]).strip(): str(r["Contractual (Yes/No)"]).strip().lower() == "yes" for _, r in sku_df.iterrows()}

lanes = [(r["Product Name"], r["CFA"]) for _, r in req.iterrows()]
net = {(r["Product Name"], r["CFA"]): float(r["Net Requirement (kL)"]) for _, r in req.iterrows()}
active_skus = sorted({s for (s, c) in lanes})

hub_ss = {(s, h): 0.0 for s in active_skus for h in HUBS}
for _, r in hub_norms.iterrows():
    if (r["Product Name"], r["Hub"]) in hub_ss:
        hub_ss[(r["Product Name"], r["Hub"])] = float(r["Hub Safety Stock (kL)"])
hub_open = {(s, h): 0.0 for s in active_skus for h in HUBS}
for _, r in oph.iterrows():
    k = (str(r["Product Name"]).strip(), str(r["Hub"]).strip())
    if k in hub_open:
        hub_open[k] = float(r["Opening Inventory (kL)"])

ub = {}
for s in active_skus:
    tot = sum(net[(ss, c)] for (ss, c) in lanes if ss == s) + sum(hub_ss[(s, h)] for h in HUBS)
    ub[s] = max(1.0, np.ceil(tot * 1.3 / BATCH))

lane_ix = {ln: i for i, ln in enumerate(lanes)}
t0 = time.time()
prob = pulp.LpProblem("Levisol_Plan", pulp.LpMinimize)
batch = {(s, p): pulp.LpVariable(f"b_{s}_{p}", 0, ub[s], "Integer") for s in active_skus for p in PLANTS}
y = pulp.LpVariable.dicts("y", (active_skus, PLANTS, HUBS), lowBound=0)
z = pulp.LpVariable.dicts("z", (range(len(lanes)), HUBS), lowBound=0)
u = pulp.LpVariable.dicts("u", range(len(lanes)), lowBound=0)
hend = pulp.LpVariable.dicts("he", (active_skus, HUBS), lowBound=0)
short = pulp.LpVariable.dicts("sh", (active_skus, HUBS), lowBound=0)

prob += (
    pulp.lpSum(batch[(s, p)] * BATCH * prod_cost[p] for s in active_skus for p in PLANTS)
    + pulp.lpSum(y[s][p][h] * ph_rate[(p, h)] for s in active_skus for p in PLANTS for h in HUBS)
    + pulp.lpSum(z[lane_ix[(s, c)]][h] * hc_rate[(h, c)] for (s, c) in lanes for h in HUBS)
    + pulp.lpSum(u[lane_ix[(s, c)]] * sku_pen[s] * (CONTR_MULT if sku_contr.get(s) else 1.0) for (s, c) in lanes)
    + pulp.lpSum(short[s][h] * sku_pen[s] * HUB_SHORT_FACTOR for s in active_skus for h in HUBS)
)
for p in PLANTS:
    for ln in CAP_LINES:
        on_line = [s for s in active_skus if sku_line[s] == ln]
        if on_line:
            prob += pulp.lpSum(batch[(s, p)] * BATCH for s in on_line) <= cap[(p, ln)]
for s in active_skus:
    for p in PLANTS:
        prob += pulp.lpSum(y[s][p][h] for h in HUBS) == batch[(s, p)] * BATCH
for s in active_skus:
    for h in HUBS:
        outb = pulp.lpSum(z[lane_ix[(ss, c)]][h] for (ss, c) in lanes if ss == s)
        prob += hub_open[(s, h)] + pulp.lpSum(y[s][p][h] for p in PLANTS) - outb == hend[s][h]
        prob += hend[s][h] + short[s][h] >= hub_ss[(s, h)]
for (s, c) in lanes:
    prob += pulp.lpSum(z[lane_ix[(s, c)]][h] for h in HUBS) + u[lane_ix[(s, c)]] == net[(s, c)]
if FORCE_CONTR:
    for (s, c) in lanes:
        if sku_contr.get(s):
            prob += u[lane_ix[(s, c)]] == 0

status = prob.solve(pulp.PULP_CBC_CMD(msg=False, timeLimit=TIME_LIMIT))
solve_secs = time.time() - t0
status_name = pulp.LpStatus[prob.status]
if status_name not in ("Optimal", "Not Solved"):
    if FORCE_CONTR:
        fail(f"Solver status: {status_name}. Most likely cause: 'Force contractual supply = Yes' with capacity too "
             f"low to meet all contractual demand. Set it back to No (penalty protection) or raise capacity.")
    fail(f"Solver returned status '{status_name}'. Check for extreme input values (e.g. all capacities zero).")
total_cost = pulp.value(prob.objective)
log("INFO", f"Solver finished in {solve_secs:.0f}s, status {status_name}, total cost Rs {total_cost:,.0f}.")

# ------------------------------------------------------------------ 5. OUTPUTS
print("STEP 5/5  Writing outputs...")
def v(x):
    return 0.0 if x.varValue is None else max(0.0, x.varValue)

prod_rows = [{"Product Name": s, "Plant": p, "Capacity Line": sku_line[s],
              "Batches (25kL)": int(round(v(batch[(s, p)]))),
              "Production (kL)": v(batch[(s, p)]) * BATCH,
              "Cost (Rs)": v(batch[(s, p)]) * BATCH * prod_cost[p]}
             for s in active_skus for p in PLANTS if v(batch[(s, p)]) > 1e-6]
prod_plan = pd.DataFrame(prod_rows) if prod_rows else pd.DataFrame(columns=["Product Name", "Plant", "Capacity Line", "Batches (25kL)", "Production (kL)", "Cost (Rs)"])

ph_rows = [{"Product Name": s, "From Plant": p, "To Hub": h, "Qty (kL)": v(y[s][p][h]),
            "Cost (Rs)": v(y[s][p][h]) * ph_rate[(p, h)]}
           for s in active_skus for p in PLANTS for h in HUBS if v(y[s][p][h]) > 1e-6]
ph_plan = pd.DataFrame(ph_rows) if ph_rows else pd.DataFrame(columns=["Product Name", "From Plant", "To Hub", "Qty (kL)", "Cost (Rs)"])

hc_rows = [{"Product Name": s, "From Hub": h, "To CFA": c, "Qty (kL)": v(z[lane_ix[(s, c)]][h]),
            "Cost (Rs)": v(z[lane_ix[(s, c)]][h]) * hc_rate[(h, c)]}
           for (s, c) in lanes for h in HUBS if v(z[lane_ix[(s, c)]][h]) > 1e-6]
hc_plan = pd.DataFrame(hc_rows) if hc_rows else pd.DataFrame(columns=["Product Name", "From Hub", "To CFA", "Qty (kL)", "Cost (Rs)"])

hub_pos = [{"Product Name": s, "Hub": h, "Opening (kL)": hub_open[(s, h)],
            "SS Target (kL)": hub_ss[(s, h)], "Ending Stock (kL)": v(hend[s][h]),
            "Shortfall (kL)": v(short[s][h]),
            "Shortfall Penalty (Rs)": v(short[s][h]) * sku_pen[s] * HUB_SHORT_FACTOR}
           for s in active_skus for h in HUBS
           if hub_ss[(s, h)] > 1e-6 or v(hend[s][h]) > 1e-6 or v(short[s][h]) > 1e-6]
hub_pos = pd.DataFrame(hub_pos) if hub_pos else pd.DataFrame(columns=["Product Name", "Hub", "Opening (kL)", "SS Target (kL)", "Ending Stock (kL)", "Shortfall (kL)", "Shortfall Penalty (Rs)"])

unmet_rows = [{"Product Name": s, "CFA": c, "Tier": sku_tier.get(s, "?"),
               "Contractual": "Yes" if sku_contr.get(s) else "No",
               "Requirement (kL)": net[(s, c)], "Unmet (kL)": v(u[lane_ix[(s, c)]]),
               "Penalty (Rs)": v(u[lane_ix[(s, c)]]) * sku_pen[s] * (CONTR_MULT if sku_contr.get(s) else 1.0)}
              for (s, c) in lanes if v(u[lane_ix[(s, c)]]) > 1e-6]
unmet = pd.DataFrame(unmet_rows).sort_values("Penalty (Rs)", ascending=False) if unmet_rows else \
    pd.DataFrame([{"Result": "All net requirements fully supplied - nothing unmet."}])

c_prod = prod_plan["Cost (Rs)"].sum() if len(prod_plan) and "Cost (Rs)" in prod_plan else 0.0
c_ph = ph_plan["Cost (Rs)"].sum() if len(ph_plan) and "Cost (Rs)" in ph_plan else 0.0
c_hc = hc_plan["Cost (Rs)"].sum() if len(hc_plan) and "Cost (Rs)" in hc_plan else 0.0
c_unmet = sum(r["Penalty (Rs)"] for r in unmet_rows)
c_short = hub_pos["Shortfall Penalty (Rs)"].sum() if len(hub_pos) and "Shortfall Penalty (Rs)" in hub_pos else 0.0
cost_summary = pd.DataFrame([
    {"Cost Component": "Production", "Rs": c_prod},
    {"Cost Component": "Transport Plant->Hub", "Rs": c_ph},
    {"Cost Component": "Transport Hub->CFA", "Rs": c_hc},
    {"Cost Component": "Penalty - unmet demand", "Rs": c_unmet},
    {"Cost Component": "Penalty - hub buffer shortfall", "Rs": c_short},
    {"Cost Component": "TOTAL", "Rs": c_prod + c_ph + c_hc + c_unmet + c_short},
])

tot_req = sum(net.values())
tot_unmet = sum(r["Unmet (kL)"] for r in unmet_rows)
fill = (tot_req - tot_unmet) / tot_req if tot_req > 0 else 1.0
dash_rows = [
    {"KPI": "Total cost (Rs)", "Value": f"{total_cost:,.0f}"},
    {"KPI": "Total requirement (kL)", "Value": f"{tot_req:,.1f}"},
    {"KPI": "Supplied (kL)", "Value": f"{tot_req - tot_unmet:,.1f}"},
    {"KPI": "Unmet (kL)", "Value": f"{tot_unmet:,.1f}"},
    {"KPI": "Fill rate", "Value": f"{fill:.1%}"},
    {"KPI": "Total production (kL)", "Value": f"{prod_plan['Production (kL)'].sum() if len(prod_plan) else 0:,.0f}"},
    {"KPI": "Solver status / time", "Value": f"{status_name} / {solve_secs:.0f}s"},
    {"KPI": "Requirement basis", "Value": REQ_BASIS},
    {"KPI": "Contractual unmet (kL)", "Value": f"{sum(r['Unmet (kL)'] for r in unmet_rows if r['Contractual'] == 'Yes'):,.2f}"},
]
for t in sorted(tier_fill):
    t_req = sum(net[(s, c)] for (s, c) in lanes if sku_tier.get(s) == t)
    t_un = sum(r["Unmet (kL)"] for r in unmet_rows if r["Tier"] == t)
    if t_req > 0:
        dash_rows.append({"KPI": f"Tier {t} fill rate", "Value": f"{(t_req - t_un) / t_req:.1%}"})
dashboard = pd.DataFrame(dash_rows)

util = pd.DataFrame([{"Plant": p, "Capacity Line": ln, "Capacity (kL)": cap[(p, ln)],
                      "Planned (kL)": sum(v(batch[(s, p)]) * BATCH for s in active_skus if sku_line[s] == ln),
                      "Utilisation": (sum(v(batch[(s, p)]) * BATCH for s in active_skus if sku_line[s] == ln) / cap[(p, ln)])
                      if cap[(p, ln)] > 0 else np.nan}
                     for p in PLANTS for ln in CAP_LINES])

norms_out = norms[["Product Name", "CFA", "Tier", "mu_d", "sd_d", "LT", "sLT", "Z",
                    "Safety Stock (kL)", "Reorder Point (kL)", "Days of Cover"]].rename(columns={
    "mu_d": "Avg Daily Demand (kL)", "sd_d": "Daily Demand Std Dev (kL)",
    "LT": "Lead Time (days)", "sLT": "Lead Time Std Dev (days)", "Z": "Z-score"})

req_out = req[["Product Name", "CFA", "Tier", "Jan-26 Forecast (kL)", buffer_col,
               "Opening Inventory (kL)", "Net Requirement (kL)"]]

if tot_unmet > 1e-6:
    log("WARN", f"{tot_unmet:,.1f} kL of demand could NOT be supplied within capacity - see 'Unmet Demand' sheet "
                 f"for exactly what, where, and the penalty cost.")
tot_short_kl = hub_pos["Shortfall (kL)"].sum() if len(hub_pos) and "Shortfall (kL)" in hub_pos else 0.0
if tot_short_kl > 1e-6:
    log("WARN", f"Hub safety-stock targets are {tot_short_kl:,.1f} kL short of target in aggregate - see "
                 f"'Hub Stock Position' sheet.")

target = write_output([
    ("Dashboard", dashboard),
    ("Cost Summary", cost_summary),
    ("Production Plan", prod_plan),
    ("Capacity Utilisation", util),
    ("Routing Plant-Hub", ph_plan),
    ("Routing Hub-CFA", hc_plan),
    ("Hub Stock Position", hub_pos),
    ("Unmet Demand", unmet),
    ("Inventory Norms (CFA)", norms_out),
    ("Inventory Norms (Hub)", hub_norms),
    ("Net Requirement", req_out),
])
print(f"\n  DONE. Total cost Rs {total_cost:,.0f} | fill rate {fill:.1%} | unmet {tot_unmet:,.1f} kL")
