"""
One-time (re)generator for the planner's INPUTS workbook.
Reads the original case data file (read-only) and produces a clean, flat, planner-editable
workbook with no merged captions, no footnotes inside data grids, and consistent key names.
Run again at any time to reset the inputs to the original case data.
"""
import os
import re
import sys
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TOOL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(os.path.dirname(TOOL_DIR), "SC Supporting Data.xlsx")
OUT = os.path.join(TOOL_DIR, "Levisol Planning Tool - INPUTS.xlsx")

CAP_LINES = ["<=1.5 LT", "3-5 LT", "7-20 LT", "50 LT", "180-210 LT"]
MONTHS = ["Jul-25", "Aug-25", "Sep-25", "Oct-25", "Nov-25", "Dec-25"]

def parse_capacity_line(pack_size):
    m = re.search(r'X\s*(\d+(?:\.\d+)?)\s*(ML|LT|KG)', str(pack_size).upper())
    v, unit = float(m.group(1)), m.group(2)
    litres = v / 1000 if unit == "ML" else v
    if litres <= 1.5: return "<=1.5 LT"
    if litres <= 5: return "3-5 LT"
    if litres <= 20: return "7-20 LT"
    if litres <= 55: return "50 LT"
    return "180-210 LT"

def main():
    if not os.path.exists(SRC):
        sys.exit(f"ERROR: cannot find source data file at {SRC}")

    # ---- read original exhibits ----
    a = pd.read_excel(SRC, sheet_name="A - Plants & Production", header=2).dropna(subset=["Production Cost (₹/kl)"])
    b = pd.read_excel(SRC, sheet_name="B - Plant-Hub Transport", header=2).dropna(subset=["To Mother Hub West (MHW)"])
    c = pd.read_excel(SRC, sheet_name="C -Hub-CFA Transport", header=2).dropna(subset=["Region"])
    d = pd.read_excel(SRC, sheet_name="D -SKU Portfolio+Penalty matrix", header=2).dropna(subset=["Pack size"])
    f = pd.read_excel(SRC, sheet_name="F - Service Levels", header=2).dropna(subset=["Volume contribution (%)"])
    g = pd.read_excel(SRC, sheet_name="G - Sales History", header=2).dropna(subset=["CFA"])
    e = pd.read_excel(SRC, sheet_name="E - Source + LT data", header=2).dropna(subset=["CFA"])
    i_all = pd.read_excel(SRC, sheet_name="I - Expected opening Inventory", header=3).dropna(subset=["CFA"])
    j = pd.read_excel(SRC, sheet_name="J - Jan Forecast", header=3).dropna(subset=["CFA"])

    # ---- Plants & Capacity ----
    plants = pd.DataFrame({
        "Plant Code": a["Plant Code"],
        "Location": a["Location"],
        "Capacity <=1.5 LT (kL)": a["Line Capacity \n<=1.5 LT (kl / month)"],
        "Capacity 3-5 LT (kL)": a["Line Capacity \n3- 5 LT (kl / month)"],
        "Capacity 7-20 LT (kL)": a["Line Capacity \n7- 20 LT (kl / month)"],
        "Capacity 50 LT (kL)": a["Line Capacity \n50 LT (kl / month)"],
        "Capacity 180-210 LT (kL)": a["Line Capacity \n180- 210LT (kl / month)"],
        "Production Cost (Rs per kL)": a["Production Cost (₹/kl)"],
    })

    # ---- Transport ----
    tp_ph = pd.DataFrame({
        "From Plant Code": [ {"Mumbai": "BOM", "Ahmedabad": "AHM", "Kolkata": "KOL"}[x] for x in b["From Plant"] ],
        "To MHW (Rs per kL)": b["To Mother Hub West (MHW)"],
        "To MHE (Rs per kL)": b["To Mother Hub East (MHE)"],
    })
    tp_hc = pd.DataFrame({
        "CFA": c["CFA"].str.strip() + " CFA",
        "Region": c["Region"],
        "From MHW (Rs per kL)": c["From Mother Hub West (MHW)"],
        "From MHE (Rs per kL)": c["From Mother Hub East (MHE)"],
    })

    # ---- SKU Master ----
    sku = pd.DataFrame({
        "Product Name": d["Product Name"],
        "Pack size": d["Pack size"],
        "Capacity Line": d["Pack size"].apply(parse_capacity_line),
        "Penalty Cost (Rs per kL)": d["Penalty cost (per kL)"],
        "Contractual (Yes/No)": np.where(d["Contractual?"].astype(str).str.contains("YES"), "Yes", "No"),
    })

    # ---- Service levels / tiers ----
    tiers = pd.DataFrame({
        "Tier": f["Tier"],
        "Volume Share": f["Volume contribution (%)"],
        "Target Fill Rate": f["Target Fill Rate (%)"].astype(str).str.rstrip("%").astype(float) / 100.0,
    })

    # ---- Demand / history / lead times / opening ----
    month_cols_src = [m + " (in kL)" for m in MONTHS]
    hist = g[["Product Name", "CFA"] + month_cols_src].copy()
    hist.columns = ["Product Name", "CFA"] + [m + " (kL)" for m in MONTHS]

    lt = pd.DataFrame({
        "Product Name": e["Product Name"],
        "CFA": e["CFA"],
        "Source Hub": e["Source"].map({"East": "MHE", "Rest of India": "MHW"}),
        "Plant to Hub LT (days)": e["LT (Plant to Hub)(in  days)"],
        "Hub to CFA LT (days)": e["LT (Hub to CFA ) (in  days)"],
        "Production LT (days)": e["Production lead time (in  days)"],
        "Production Variability (days)": e["Production variability (in  days)"],
        "Transit Variability (days)": e["Transit lead variability (in  days)"],
    })

    demand = j[["Product Name", "CFA", "Jan -2026 (in kL)"]].copy()
    demand.columns = ["Product Name", "CFA", "Jan-26 Forecast (kL)"]

    op_cfa = i_all[~i_all["CFA"].str.contains("Mother Hub")][["Product Name", "CFA", "Jan -2026 (in kL)"]].copy()
    op_cfa.columns = ["Product Name", "CFA", "Opening Inventory (kL)"]
    op_hub_raw = i_all[i_all["CFA"].str.contains("Mother Hub")].copy()
    op_hub = pd.DataFrame({
        "Product Name": op_hub_raw["Product Name"],
        "Hub": op_hub_raw["CFA"].map({"Mother Hub West": "MHW", "Mother Hub East": "MHE"}),
        "Opening Inventory (kL)": op_hub_raw["Jan -2026 (in kL)"],
    })

    # ---- Settings ----
    settings = pd.DataFrame([
        ["Solver time limit (seconds)", 150, "How long the optimiser may search. 60 = quick what-if, 150 = normal run, 600 = final run before submission."],
        ["Contractual penalty multiplier", 5, "Unmet demand on contractual SKUs is charged at this multiple of the SKU's penalty cost."],
        ["Hub shortfall penalty factor", 0.5, "Ending a month below a hub's safety-stock target is charged at this fraction of the SKU's penalty cost per kL."],
        ["Batch size (kL)", 25, "All production quantities are integer multiples of this. From the case: 25."],
        ["Hub service level", 0.98, "Service level used for hub safety-stock targets (case instruction: 98% for all grades)."],
        ["Requirement basis (ROP or SS)", "ROP", "ROP: replenish to reorder point (primary, per assignment). SS: replenish to safety stock only (leaner alternative)."],
        ["Working days per month", 30, "From the case: 30 working days."],
        ["Force contractual supply (Yes/No)", "No", "Yes = contractual SKU demand MUST be fully met (hard rule). No = protected by the penalty multiplier instead."],
    ], columns=["Setting", "Value", "What it means"])

    # ---- Hub SS override (empty template) ----
    override = pd.DataFrame(columns=["Product Name", "Hub (MHW/MHE)", "Override Safety Stock (kL)"])

    # ---- README ----
    readme_rows = [
        ["LEVISOL MONTHLY PLANNING TOOL - HOW TO USE"],
        [""],
        ["This workbook is the BASE INPUT DATA for the Levisol web planning tool."],
        [""],
        ["NORMAL USE: launch the web app (double-click 'RUN WEB APP.bat', or the deployed URL) and"],
        ["edit everything directly in the browser's Inputs tab - you rarely need to open this file."],
        [""],
        ["This workbook is loaded as the app's starting data. If you prefer to prepare inputs in"],
        ["Excel: edit the yellow sheets here, save and CLOSE the file, then launch the web app -"],
        ["it picks up this file on start."],
        [""],
        ["IF SOMETHING IS WRONG WITH THE INPUTS: the tool will NOT crash. It shows a Run Log"],
        ["explaining exactly which cell/sheet has the problem."],
        [""],
        ["TYPICAL RUN TIME: ~1 minute at the 60-second solver setting (within a few % of optimal),"],
        ["about 2.5 minutes at 150 seconds; use 600 seconds for a final run."],
        [""],
        ["SHEET GUIDE"],
        ["  Settings ................. tool behaviour (time limit, penalty rules, requirement basis)"],
        ["  Jan Demand Forecast ...... the demand the plan is built for  (MOST COMMONLY EDITED)"],
        ["  Plants & Capacity ........ per-line monthly capacity (kL) and production cost per plant"],
        ["  Transport Plant-Hub ...... freight Rs/kL from each plant to each hub"],
        ["  Transport Hub-CFA ........ freight Rs/kL from each hub to each CFA"],
        ["  SKU Master ............... pack size, capacity line, penalty cost, contractual flag"],
        ["  Service Levels (Tiers) ... tier definitions and target fill rates"],
        ["  Sales History ............ last 6 months' sales, used to compute demand variability & tiers"],
        ["  Lead Times ............... replenishment lead times & variability per SKU-CFA"],
        ["  Opening Inventory CFA .... stock at each CFA at the start of the month"],
        ["  Opening Inventory Hub .... stock at each hub at the start of the month"],
        ["  Hub SS Override .......... optional: force a specific hub safety-stock target for a SKU"],
        [""],
        ["To reset all inputs to the original case data: use 'Reset all inputs' in the web app,"],
        ["or run engine\\build_inputs_workbook.py to regenerate this file."],
    ]
    readme = pd.DataFrame(readme_rows, columns=[" "])

    sheets = {
        "READ ME FIRST": (readme, "1F6FC0"),
        "Settings": (settings, "FFD34D"),
        "Jan Demand Forecast": (demand, "FFD34D"),
        "Plants & Capacity": (plants, "FFD34D"),
        "Transport Plant-Hub": (tp_ph, "FFD34D"),
        "Transport Hub-CFA": (tp_hc, "FFD34D"),
        "SKU Master": (sku, "FFD34D"),
        "Service Levels (Tiers)": (tiers, "FFD34D"),
        "Sales History": (hist, "FFD34D"),
        "Lead Times": (lt, "FFD34D"),
        "Opening Inventory CFA": (op_cfa, "FFD34D"),
        "Opening Inventory Hub": (op_hub, "FFD34D"),
        "Hub SS Override": (override, "B8E0B8"),
    }

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        for name, (df, colour) in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            ws.sheet_properties.tabColor = colour
            ws.freeze_panes = "A2"
            bold = Font(bold=True)
            fill = PatternFill("solid", fgColor="DDDDDD")
            for cell in ws[1]:
                cell.font = bold
                if name != "READ ME FIRST":
                    cell.fill = fill
            for idx, col in enumerate(df.columns, start=1):
                width = max(14, min(60, len(str(col)) + 4))
                if name == "READ ME FIRST":
                    width = 110
                ws.column_dimensions[get_column_letter(idx)].width = width
            if name == "READ ME FIRST":
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=False)

    print(f"Inputs workbook written to: {OUT}")
    print(f"Sheets: {list(sheets.keys())}")

if __name__ == "__main__":
    main()
